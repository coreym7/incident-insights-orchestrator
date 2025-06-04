from datetime import datetime
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
import re
from collections import defaultdict
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side

# Read CSV to DataFrame
def read_csv_to_dataframe(file_path, column_mapping):
    """
    Reads a CSV file into a DataFrame and standardizes column names using a mapping dictionary.
    """
    df = pd.read_csv(file_path)
    df.rename(columns=column_mapping, inplace=True)
   
    return df

# Convert DataFrame to List of Dictionaries
def dataframe_to_dict_list(dataframe):
    """
    Converts a DataFrame to a list of dictionaries for further processing.
    """
    return dataframe.to_dict(orient='records')

def process_input_file(input_file, column_mapping):
    """
    Reads the input file and returns the data as a list of dictionaries.
    Args:
        input_file (str): Path to the input CSV file.
        column_mapping (dict): Dictionary mapping original column names to standardized names.
    Returns:
        list: A list of dictionaries containing the processed data.
    """
    # Step 1: Read the CSV into a DataFrame
    df = read_csv_to_dataframe(input_file, column_mapping)
    
    # Step 2: Convert the DataFrame to a list of dictionaries
    return dataframe_to_dict_list(df)

# Calculate Summary Metrics
def calculate_summary_metrics(data):
    """
    Calculate summary metrics and return a dictionary of all results.
    """
    # Placeholder for all metric calculations
    metrics = {}
    metrics['incidents_by_grade'] = count_by_grade(data)
    metrics['incidents_by_location'] = count_by_location(data)
    metrics['incidents_by_hour'] = count_by_hour(data)
    metrics['incidents_by_date'] = count_by_date(data)
    metrics['incidents_by_subtype'] = count_by_subtype(data)
    metrics['top_students'] = top_students(data, top_n=15) #top students may change as time of year changes
    metrics['top_authors'] = top_authors(data, top_n=10)
    metrics['incidents_by_loc_hour'] = hourly_location(data)
    return metrics

def count_by_grade(data):
    """
    Counts the number of incidents per grade level.
    Args:
        data (list of dict): The discipline log data.
    Returns:
        list: A list of dictionaries with 'Grade' and 'Count' as keys.
    """
    grade_counts = {}
    for row in data:
        grade = row.get('grade_level', 'Unknown')  # Handle missing grades as 'Unknown'
        grade_counts[grade] = grade_counts.get(grade, 0) + 1

    # Convert the dictionary into a list of dictionaries
    return [{"Grade": grade, "Count": count} for grade, count in grade_counts.items()]

def count_by_location(data):
    """
    Counts the number of incidents per location.
    Args:
        data (list of dict): The discipline log data.
    Returns:
        list: A list of dictionaries with 'Location' and 'Count' as keys.
    """
    location_counts = {}
    for row in data:
        location = row.get('incident_location', 'Unknown')  # Handle missing locations as 'Unknown'
        location_counts[location] = location_counts.get(location, 0) + 1

    # Convert the dictionary into a list of dictionaries
    return [{"Location": location, "Count": count} for location, count in location_counts.items()]

def count_by_hour(data):
    """
    Counts the number of incidents per hour of the day and formats hours as 12-hour time with AM/PM suffix.
    Args:
        data (list of dict): The discipline log data.
    Returns:
        list: A list of dictionaries with 'Hour' and 'Count' as keys, where 'Hour' is in readable format.
    """
    hour_counts = {}
    for row in data:
        time_str = row.get('incident_time', None)  # Extract the time string
        if time_str and isinstance(time_str, str):  # Ensure time_str is a valid string
            try:
                # Try parsing with seconds
                time_obj = datetime.strptime(time_str.strip(), "%I:%M:%S %p")
            except ValueError:
                try:
                    # Fallback to parsing without seconds
                    time_obj = datetime.strptime(time_str.strip(), "%I:%M %p")
                except ValueError:
                    try:
                        # Handle times without AM/PM (e.g., '2:45') by assuming 24-hour format
                        time_obj = datetime.strptime(time_str.strip(), "%H:%M")
                    except ValueError as e:
                        # Debug: Print the error for invalid times
                        print(f"Failed to parse time: '{time_str}' -> {e}")
                        hour = "Unknown"
                        hour_counts[hour] = hour_counts.get(hour, 0) + 1
                        continue

            # Format to '2p', '3p', etc., stripping leading zeros
            hour = time_obj.strftime("%I%p").lower().lstrip("0")
        else:
            # Handle missing, non-string, or empty time
            hour = "Unknown"

        # Update the counts
        hour_counts[hour] = hour_counts.get(hour, 0) + 1

    # Ensure "Unknown" is always in hour_counts
    hour_counts["Unknown"] = hour_counts.get("Unknown", 0)

    # Sort hours and add "Unknown" at the end
    sorted_hours = sorted(
        [h for h in hour_counts.keys() if h != "Unknown"],
        key=lambda x: datetime.strptime(x, "%I%p")
    )
    sorted_hours.append("Unknown")

    return [{"Hour": hour, "Count": hour_counts[hour]} for hour in sorted_hours]


def count_by_date(data):
    """
    Counts the number of incidents per specific date and calculates the average by day of the week.
    """
    from collections import defaultdict
    import pandas as pd

    # Step 1: Initialize counters
    date_counts = defaultdict(int)  # Total incidents per date
    day_totals = defaultdict(int)   # Total incidents per day of week
    day_counts = defaultdict(int)   # Count of how many times each day occurs

    for row in data:
        date_str = row.get('incident_date', None)
        if date_str and isinstance(date_str, str):
            #print(f"Raw date string: {date_str}")  # Debugging statement

            try:
                # Parse date using pd.to_datetime for robustness
                date_obj = pd.to_datetime(date_str, format="%m/%d/%Y", errors="coerce")
                if pd.isnull(date_obj):  # Retry with alternative format
                    date_obj = pd.to_datetime(date_str, format="%m-%d-%Y", errors="coerce")
                
                if pd.isnull(date_obj):  # If still invalid, log it
                    print(f"Invalid date format: {date_str}")
                    continue

                formatted_date = date_obj.strftime("%m/%d/%Y")
                #print(f"Parsed date: {formatted_date}")  # Debugging statement

                date_counts[formatted_date] += 1

                # Track day of the week (e.g., Monday, Tuesday)
                day_of_week = date_obj.strftime("%A")
                day_totals[day_of_week] += 1
                
            except Exception as e:
                print(f"Error parsing date '{date_str}': {e}")  # Debugging statement
        else:
            print(f"Invalid or missing date in row: {row}")  # Debugging statement

    # Calculate unique date counts for each day of the week
    for date, count in date_counts.items():
        try:
            date_obj = pd.to_datetime(date, format="%m/%d/%Y")
            day_of_week = date_obj.strftime("%A")
            day_counts[day_of_week] += 1  # Increment by 1 for each unique date
        except Exception as e:
            print(f"Error re-parsing date '{date}': {e}")  # Debugging statement


    # Step 2: Calculate averages per day of the week
    day_of_week_avg = []
    for day, total in day_totals.items():
        avg = total / day_counts[day]  # Divide total incidents by number of that day
        day_of_week_avg.append({"Day of Week": day, "Average Incidents": round(avg, 2)})

    # Sort the days of the week explicitly
    week_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    day_of_week_avg.sort(key=lambda x: week_order.index(x["Day of Week"]))

    return {
        "date_counts": [{"Date": date, "Count": count} for date, count in date_counts.items()],
        "day_of_week_avg": day_of_week_avg
    }




def count_by_subtype(data):
    """
    Counts the number of incidents per subtype.
    Args:
        data (list of dict): The discipline log data.
    Returns:
        list: A list of dictionaries with 'Subtype' and 'Count' as keys.
    """
    subtype_counts = {}
    for row in data:
        subtype = row.get('subtype_name', 'Unknown')  # Handle missing subtype as 'Unknown'
        subtype_counts[subtype] = subtype_counts.get(subtype, 0) + 1

    # Convert the dictionary into a list of dictionaries
    return [{"Subtype": subtype, "Count": count} for subtype, count in subtype_counts.items()]


def top_students(data, top_n=10):
    """
    Identifies the top students by number of incidents.
    Args:
        data (list of dict): The discipline log data.
        top_n (int): Number of top students to include.
    Returns:
        list: A list of dictionaries with 'Student' and 'Incidents' as keys.
    """
    student_counts = {}
    for row in data:
        student = row.get('student_name', 'Unknown')  # Handle missing student names as 'Unknown'
        student_counts[student] = student_counts.get(student, 0) + 1

    # Sort students by count in descending order and take the top `n`
    sorted_students = sorted(student_counts.items(), key=lambda x: x[1], reverse=True)[:top_n]

    # Convert the sorted list into a list of dictionaries
    return [{"Student": student, "Incidents": count} for student, count in sorted_students]


def top_authors(data, top_n=10):
    """
    Identifies the top authors by number of logs entered.
    Args:
        data (list of dict): The discipline log data.
        top_n (int): Number of top authors to include.
    Returns:
        list: A list of dictionaries with 'Author' and 'Logs' as keys.
    """
    author_counts = {}
    for row in data:
        author = row.get('entry_author', 'Unknown')  # Handle missing authors as 'Unknown'
        author_counts[author] = author_counts.get(author, 0) + 1

    # Sort authors by count in descending order and take the top `n`
    sorted_authors = sorted(author_counts.items(), key=lambda x: x[1], reverse=True)[:top_n]

    # Convert the sorted list into a list of dictionaries
    return [{"Author": author, "Logs": count} for author, count in sorted_authors]



# Load CSV to Workbook and Delete CSV
def csv_to_excel_workbook(csv_paths, workbook_path):
    """
    Load all CSV files into an Excel workbook as separate sheets, then delete the CSV files.
    """
    with pd.ExcelWriter(workbook_path) as writer:
        for key, csv_path in csv_paths.items():
            df = pd.read_csv(csv_path)
            df.to_excel(writer, sheet_name=key, index=False)
    # Clean up temporary CSV files
    #for csv_path in csv_paths.values():
        #os.remove(csv_path)
    
def consolidate_metrics_to_csv(metrics, output_path):
    """
    Consolidates all metrics into a single CSV file.
    Args:
        metrics (dict): A dictionary of metrics where each value is a list of dictionaries.
        output_path (str): The path to the consolidated CSV file.
    """
    with open(output_path, 'w') as file:
        for metric_name, metric_data in metrics.items():
            # Write headers explicitly for each metric
            if metric_data:
                headers = metric_data[0].keys()  # Get column names from the first row
                file.write(",".join(headers) + "\n")  # Write column headers
                
                for row in metric_data:
                    file.write(",".join(map(str, row.values())) + "\n")  # Write rows
                
                # Add a blank line between metrics
                file.write("\n")

def hourly_location(data):
    """
    Calculates the breakdown of incidents by hour and location.
    Args:
        data (list of dict): The discipline log data.
    Returns:
        list: A list of dictionaries with 'Hour', 'Location', and 'Count' as keys.
    """
    from collections import defaultdict

    # Initialize nested dictionary for hourly location counts
    hourly_location_counts = defaultdict(lambda: defaultdict(int))

    # Populate the counts
    for row in data:
        # Get hour and location
        incident_time = row.get("incident_time", None)
        location = row.get("incident_location", "Unknown")

        # Cast incident_time to a string if it's not None
        if incident_time is not None:
            incident_time = str(incident_time).strip()

        # Parse the time or set as Unknown
        if incident_time:
            try:
                time_obj = datetime.strptime(incident_time, "%I:%M %p")
                hour = time_obj.strftime("%I%p").lower().lstrip("0")  # e.g., '10am', '2pm'
            except ValueError:
                hour = "Unknown"
        else:
            hour = "Unknown"

        # Update count for the location at the given hour
        hourly_location_counts[hour][location] += 1

    # Convert to a list of dictionaries for compatibility with metrics structure
    breakdown_rows = []
    for hour, locations in sorted(
        hourly_location_counts.items(),
        key=lambda x: datetime.strptime(x[0], "%I%p") if x[0] != "Unknown" else datetime.max,
    ):
        for location, count in locations.items():
            breakdown_rows.append({"Hour": hour, "Location": location, "Count": count})

    return breakdown_rows


def write_metrics_to_workbook(metrics, workbook_path):
    """
    Writes all metrics to an Excel workbook with section headers, including compact layout adjustments.
    Args:
        metrics (dict): A dictionary of metrics where each value is a list of dictionaries or nested data.
        workbook_path (str): The path to save the Excel workbook.
    """
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    current_row = 1
    current_col = 1  # Start in column A by default

    ###########################new#################################
    # Define border styles
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    ###########################new#################################

    for metric_name, metric_data in metrics.items():
        # Reset position for specific metrics
        
        if metric_name == "incidents_by_grade":
            current_row = 1
            current_col = 1  # Column A
        elif metric_name == "incidents_by_loc_hour":
            current_row = 1
            current_col = 4 
        elif metric_name == "incidents_by_date":
            current_row = 1
            current_col = 8  # Column D
        elif metric_name == "incidents_by_subtype":
            current_row = 1
            current_col = 11  # Column G
        elif metric_name == "top_students":
            current_row = 1
            current_col = 14 # Column J
        else:
            current_row += 2  # Add spacing between sections

        # Add the metric name as a bold header
        ws.cell(row=current_row, column=current_col, value=metric_name).font = Font(bold=True)
        current_row += 1

        if not metric_data:
            # Write a placeholder if the metric has no data
            ws.cell(row=current_row, column=current_col, value="No Data Available").font = Font(italic=True)
            current_row += 2
            continue

        if isinstance(metric_data, dict):
            # Handle nested data for incidents_by_date
            if metric_name == "incidents_by_date":
                # Force specific order: day_of_week_avg first, then date_counts
                ordered_sub_metrics = ["day_of_week_avg", "date_counts"]
            else:
                # Use the default order for other metrics
                ordered_sub_metrics = metric_data.keys()

            for sub_metric_name in ordered_sub_metrics:
                sub_metric_data = metric_data.get(sub_metric_name, [])
                ws.cell(row=current_row, column=current_col, value=sub_metric_name).font = Font(bold=True)
                current_row += 1

                if sub_metric_data:
                    # Write column headers
                    headers = list(sub_metric_data[0].keys())
                    for col_num, header in enumerate(headers, start=current_col):
                        ws.cell(row=current_row, column=col_num, value=header).font = Font(bold=True)
                    current_row += 1

                    # Write rows of data
                    for row in sub_metric_data:
                        for col_num, value in enumerate(row.values(), start=current_col):
                            ws.cell(row=current_row, column=col_num, value=value)
                        current_row += 1

                # Add a blank row between sections of nested data
                current_row += 1

        # Handle 'incidents_by_loc_hour' as an elif block here
        elif metric_name == "incidents_by_loc_hour":
            # Write column headers
            headers = ["Hour", "Location", "Count"]
            for col_num, header in enumerate(headers, start=current_col):
                ws.cell(row=current_row, column=col_num, value=header).font = Font(bold=True)
            current_row += 1

            # Track the start of each grouping for borders
            group_start_row = current_row
            last_hour = None

            for row in metric_data:
                hour = row["Hour"]

                # If the hour changes, apply a border around the previous group
                if last_hour is not None and hour != last_hour:
                    for group_row in range(group_start_row, current_row):  # Loop through all rows in the group
                        for col in range(current_col, current_col + len(headers)):
                            ws.cell(row=group_row, column=col).border = Border(
                                left=Side(style="thin"),
                                right=Side(style="thin")
                            )
                    # Add top and bottom borders to the first and last rows of the group
                    for col in range(current_col, current_col + len(headers)):
                        ws.cell(row=group_start_row, column=col).border = Border(
                            top=Side(style="thin"),
                            left=Side(style="thin"),
                            right=Side(style="thin")
                        )
                        ws.cell(row=current_row - 1, column=col).border = Border(
                            bottom=Side(style="thin"),
                            left=Side(style="thin"),
                            right=Side(style="thin")
                        )
                    group_start_row = current_row  # Reset the start row for the new hour group

                # Write data rows
                for col_num, value in enumerate(row.values(), start=current_col):
                    ws.cell(row=current_row, column=col_num, value=value)

                current_row += 1
                last_hour = hour

            # Apply a border to the final group
            for group_row in range(group_start_row, current_row):  # Loop through all rows in the final group
                for col in range(current_col, current_col + len(headers)):
                    ws.cell(row=group_row, column=col).border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin")
                    )
            for col in range(current_col, current_col + len(headers)):
                ws.cell(row=group_start_row, column=col).border = Border(
                    top=Side(style="thin"),
                    left=Side(style="thin"),
                    right=Side(style="thin")
                )
                ws.cell(row=current_row - 1, column=col).border = Border(
                    bottom=Side(style="thin"),
                    left=Side(style="thin"),
                    right=Side(style="thin")
                )
            current_row += 1  # Add space after the section
        

        else:
            # Handle standard metrics
            headers = list(metric_data[0].keys())
            for col_num, header in enumerate(headers, start=current_col):
                ws.cell(row=current_row, column=col_num, value=header).font = Font(bold=True)
            current_row += 1

            for row in metric_data:
                for col_num, value in enumerate(row.values(), start=current_col):
                    ws.cell(row=current_row, column=col_num, value=value).border = thin_border
                current_row += 1

        # Add a blank row between metrics
        current_row += 1

    # Save the workbook
    wb.save(workbook_path)




def ensure_output_folder(output_folder):
    """
    Ensures the output folder exists.
    Args:
        output_folder (str): The folder to create if it doesn't exist.
    """
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

def sanitize_building_name(building_name):
    """
    Sanitize the building name to ensure it can be safely used in file and folder paths.
    Args:
        building_name (str): The original building name.
    Returns:
        str: A sanitized building name.
    """
    if not isinstance(building_name, str):
        building_name = str(building_name) if building_name is not None else "Unknown Building"
    # Replace invalid characters with underscores
    return re.sub(r'[<>:"/\\|?*]', '_', building_name)

def add_hourly_location_breakdown(workbook_path, building_data):
    """
    Adds a breakdown of incident counts by hour and location as a new sheet to the workbook.
    Args:
        workbook_path (str): Path to the existing Excel workbook.
        building_data (list of dict): Log entries for the building.
    """
    from collections import defaultdict
    import pandas as pd

    # Initialize nested dictionary for hourly location counts
    hourly_location_counts = defaultdict(lambda: defaultdict(int))

    # Populate the counts
    for row in building_data:
        # Get hour and location
        incident_time = row.get("incident_time", None)
        location = row.get("incident_location", "Unknown")

        # Cast incident_time to a string if it's not None
        if incident_time is not None:
            incident_time = str(incident_time).strip()

        # Parse the time or set as Unknown
        if incident_time:
            try:
                time_obj = datetime.strptime(incident_time, "%I:%M %p")
                hour = time_obj.strftime("%I%p").lower().lstrip("0")  # e.g., '10am', '2pm'
            except ValueError:
                hour = "Unknown"
        else:
            hour = "Unknown"

        # Update count for the location at the given hour
        hourly_location_counts[hour][location] += 1

    # Convert to a DataFrame for better Excel output
    breakdown_rows = []
    for hour, locations in sorted(
        hourly_location_counts.items(),
        key=lambda x: datetime.strptime(x[0], "%I%p") if x[0] != "Unknown" else datetime.max,
    ):
        for location, count in locations.items():
            breakdown_rows.append({"Hour": hour, "Location": location, "Count": count})

    breakdown_df = pd.DataFrame(breakdown_rows)

    # Append to the workbook as a new sheet
    with pd.ExcelWriter(workbook_path, engine="openpyxl", mode="a") as writer:
        breakdown_df.to_excel(writer, sheet_name="Hourly Location Breakdown", index=False)



def generate_building_reports(input_file, column_mapping, output_folder, workbook_name_template):
    """
    Generates reports for each building based on the 'Student School' column in the data.
    Args:
        input_file (str): Path to the district-wide input CSV file.
        column_mapping (dict): Column mapping dictionary for standardizing column names.
        output_folder (str): Base output folder where building reports will be stored.
        workbook_name_template (str): Template for naming building-specific workbooks (e.g., "{Building_Name}_Report.xlsx").
    """
    # Step 1: Read and process the district-wide input data
    data = process_input_file(input_file, column_mapping)

    # Step 2: Group data by 'Student School'
    building_groups = {}
    for row in data:
        building_name = row.get('student_school', "Unknown Building")  # Handle missing values
        if building_name not in building_groups:
            building_groups[building_name] = []
        building_groups[building_name].append(row)

    # Step 3: Process data for each building
    for building_name, building_data in building_groups.items():
        safe_building_name = sanitize_building_name(building_name)
        # Create a subfolder for the building
        building_folder = os.path.join(output_folder, safe_building_name)
        ensure_output_folder(building_folder)

        # Generate a workbook name for the building
        workbook_name = workbook_name_template.format(Building_Name=safe_building_name.replace(" ", "_"))
        workbook_path = os.path.join(building_folder, workbook_name)

        # Calculate metrics and write the report for this building
        metrics = calculate_summary_metrics(building_data)
        write_metrics_to_workbook(metrics, workbook_path)

        # Append raw log entries as a new sheet
        append_raw_data_to_workbook(workbook_path, building_data)

        # Add the new hourly-location breakdown sheet
        #add_hourly_location_breakdown(workbook_path, building_data)

    print(f"Building reports generated successfully in {output_folder}")

def append_raw_data_to_workbook(workbook_path, building_data):
    """
    Appends raw log entries as a new sheet to an existing workbook.
    Args:
        workbook_path (str): Path to the existing Excel workbook.
        building_data (list of dict): Raw log entries for the building.
    """
    # Convert the building data to a DataFrame
    raw_data_df = pd.DataFrame(building_data)

    # Load the existing workbook
    with pd.ExcelWriter(workbook_path, engine="openpyxl", mode="a") as writer:
        # Add raw log entries as a new sheet
        raw_data_df.to_excel(writer, sheet_name="Detailed Log Entries", index=False)

def main(input_file, column_mapping, output_folder, workbook_name):
    """
    Main function to process the input file and generate both district-wide and per-building reports.
    """
    # Ensure the output directory exists
    ensure_output_folder(output_folder)

    # Get current date in the desired format (e.g., YYYY-MM-DD)
    current_date = datetime.now().strftime("%Y-%m-%d")

        # Step 1: Generate the district-wide report
    data = process_input_file(input_file, column_mapping)
    metrics = calculate_summary_metrics(data)
    write_metrics_to_workbook(metrics, os.path.join(output_folder, workbook_name))

    # Step 2: Generate per-building reports
    generate_building_reports(
        input_file=input_file,
        column_mapping=column_mapping,
        output_folder=output_folder,
        workbook_name_template=f"{{Building_Name}}_Report_{current_date}.xlsx"
    )


# Execute
if __name__ == "__main__":
    input_file = "discipline_logs.csv"
    column_mapping = {
    'Student Number': 'student_number',
    'Student Name': 'student_name',
    'Grade Level': 'grade_level',
    'Entry Author': 'entry_author',
    'Entry Date': 'entry_date',
    'Entry Hour': 'entry_hour',
    'Entry Minute': 'entry_minute',
    'Entry Meridiem': 'entry_meridiem',
    'Incident Date': 'incident_date',
    'Incident Time': 'incident_time',
    'Category': 'category',
    'Subject': 'subject',
    'Entry': 'entry',
    'Submitted By Teacher': 'submitted_by_teacher',
    'Log Type': 'log_type',
    'Subtype Name': 'subtype_name',
    'Consequence': 'consequence',
    'Consequence Name': 'consequence_name',
    'Incident Location': 'incident_location',
    'Student School': 'student_school'
}

    output_folder = "./output"
    workbook_name = "Discipline_Report.xlsx"

    main(input_file, column_mapping, output_folder, workbook_name)
